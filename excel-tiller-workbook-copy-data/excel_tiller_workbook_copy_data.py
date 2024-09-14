import pandas as pd
from openpyxl import load_workbook

# Request file paths from the user
source_file = input("Enter the path to the source Excel file: ").strip('"')
destination_file = input("Enter the path to the destination Excel file: ").strip('"')

# Define the columns for each table
transactions_columns = [
    'Date', 'Description', 'Category', 'Amount', 'Account',
    'Account #', 'Institution', 'Month', 'Week', 'Transaction ID', 'Account ID',
    'Check Number', 'Full Description', 'Date Added'
]
account_columns = ['Account', 'Class Override', 'Group', 'Hide']
balance_history_columns = [
    'Account', 'Account #', 'Account ID', 'Institution', 'Balance', 'Month',
    'Week', 'Type', 'Date Added'
]
categories_columns = [
    'Category', 'Group', 'Type', 'Hide From Reports'
]

# Read and filter the data
sheets = {
    "Transactions": transactions_columns,
    "Accounts": account_columns,
    "Balance History": balance_history_columns,
    "Categories": categories_columns
}

# Load the destination workbook
wb = load_workbook(destination_file)

for sheet_name, columns in sheets.items():
    # Read data from the source file
    df = pd.read_excel(source_file, sheet_name=sheet_name, usecols=columns)

    # Load the sheet from the destination workbook
    ws = wb[sheet_name]

    # Overwrite only the values in the destination sheet
    for i, row in df.iterrows():
        for j, value in enumerate(row):
            # Excel is 1-indexed, so row/col start at 1,1
            ws.cell(row=i + 2, column=j + 1, value=value)

# Save the workbook with updated data
wb.save(destination_file)

print("Data copied successfully!")
